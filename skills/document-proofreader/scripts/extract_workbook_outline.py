#!/usr/bin/env python3
"""
Extract a read-only outline from an Excel workbook for cross-document copy proofreading.

The script never modifies the input workbook. It produces JSON with:
- sheet dimensions
- candidate header rows
- detected Chinese / English / French / North America columns
- styled/commented cells that may indicate change notes
- row text for detected copy columns
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2) from exc


TEXT_LIMIT = 500
DEFAULT_HEADER_SCAN_ROWS = 30


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def truncate(value: str, limit: int = TEXT_LIMIT) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def tag_header(header: str) -> List[str]:
    """Return semantic tags for a column header while avoiding common false positives."""
    raw = header.strip()
    lower = raw.lower()
    tags: List[str] = []

    def has_token(token: str) -> bool:
        return re.search(rf"(?<![a-z0-9]){re.escape(token)}(?![a-z0-9])", lower) is not None

    if any(k in raw for k in ["中文", "原文", "源文案"]):
        tags.append("chinese_source")
    if has_token("cn") or "chinese" in lower:
        tags.append("chinese_source")

    if "英文" in raw or "english" in lower or "us english" in lower or has_token("en"):
        tags.append("english")

    if "法文" in raw or "french" in lower or "canadian french" in lower or has_token("fr") or has_token("fr-ca"):
        tags.append("french")

    # North America tags. Avoid n/a and words like name.
    if "北美" in raw or "north america" in lower or "canada" in lower or has_token("us"):
        tags.append("north_america")
    if has_token("na") and "n/a" not in lower and "name" not in lower:
        tags.append("north_america")

    # Deduplicate while preserving order.
    seen = set()
    result = []
    for tag in tags:
        if tag not in seen:
            seen.add(tag)
            result.append(tag)
    return result


def is_meaningful_fill(cell: Any) -> bool:
    fill = getattr(cell, "fill", None)
    if fill is None:
        return False
    pattern_type = getattr(fill, "patternType", None)
    if not pattern_type:
        return False
    if pattern_type == "none":
        return False
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return True
    rgb = getattr(fg, "rgb", None)
    indexed = getattr(fg, "indexed", None)
    theme = getattr(fg, "theme", None)
    return bool(rgb or indexed is not None or theme is not None)


def color_to_string(color: Any, include_theme: bool = False) -> Optional[str]:
    """Return a stable color string without triggering openpyxl descriptor errors."""
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        rgb = getattr(color, "rgb", None)
        return str(rgb) if isinstance(rgb, str) and rgb else None
    if color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        return f"indexed:{indexed}" if indexed is not None else None
    if color_type == "theme" and include_theme:
        theme = getattr(color, "theme", None)
        tint = getattr(color, "tint", None)
        return f"theme:{theme},tint:{tint}" if theme is not None else None
    return None


def font_color(cell: Any) -> Optional[str]:
    font = getattr(cell, "font", None)
    color = getattr(font, "color", None) if font else None
    # Ignore theme colors because default workbook fonts often use theme colors.
    return color_to_string(color, include_theme=False)


def fill_color(cell: Any) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    fg = getattr(fill, "fgColor", None) if fill else None
    return color_to_string(fg, include_theme=True)


def row_values(ws: Any, row_index: int) -> List[str]:
    values = []
    for col_idx in range(1, ws.max_column + 1):
        values.append(normalize_text(ws.cell(row=row_index, column=col_idx).value))
    return values


def candidate_header_rows(ws: Any, scan_rows: int) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    max_scan = min(ws.max_row or 0, scan_rows)
    for row_idx in range(1, max_scan + 1):
        values = row_values(ws, row_idx)
        nonempty = [v for v in values if v]
        hits = []
        for col_idx, value in enumerate(values, start=1):
            tags = tag_header(value)
            if tags:
                hits.append({
                    "column": get_column_letter(col_idx),
                    "index": col_idx,
                    "header": truncate(value, 120),
                    "tags": tags,
                })
        score = len(hits) * 5 + min(len(nonempty), 10)
        if hits or len(nonempty) >= 3:
            candidates.append({
                "row": row_idx,
                "score": score,
                "nonempty_count": len(nonempty),
                "language_hits": hits,
                "sample_values": [truncate(v, 120) for v in nonempty[:8]],
            })
    candidates.sort(key=lambda item: (-item["score"], item["row"]))
    return candidates[:8]


def detect_columns(ws: Any, header_row: Optional[int]) -> List[Dict[str, Any]]:
    if not header_row:
        return []
    columns = []
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_text(ws.cell(row=header_row, column=col_idx).value)
        tags = tag_header(header)
        if header or tags:
            columns.append({
                "index": col_idx,
                "letter": get_column_letter(col_idx),
                "header": truncate(header, 160),
                "tags": tags,
            })
    return columns


def extract_styled_or_commented_cells(ws: Any, max_items: int = 200) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = normalize_text(cell.value)
            comment = normalize_text(cell.comment.text) if cell.comment else ""
            fill = fill_color(cell) if is_meaningful_fill(cell) else None
            fcolor = font_color(cell)
            if comment or fill or fcolor:
                items.append({
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "column": get_column_letter(cell.column),
                    "value": truncate(value),
                    "comment": truncate(comment),
                    "fill_color": fill,
                    "font_color": fcolor,
                })
                if len(items) >= max_items:
                    return items
    return items


def relevant_column_indices(columns: Iterable[Dict[str, Any]]) -> List[int]:
    result = []
    for col in columns:
        tags = set(col.get("tags", []))
        if tags.intersection({"chinese_source", "english", "french", "north_america"}):
            result.append(int(col["index"]))
    return result


def extract_text_rows(ws: Any, header_row: Optional[int], columns: List[Dict[str, Any]], max_rows: int) -> List[Dict[str, Any]]:
    if not header_row:
        return []
    indices = relevant_column_indices(columns)
    if not indices:
        return []
    start_row = header_row + 1
    end_row = ws.max_row or 0
    if max_rows and max_rows > 0:
        end_row = min(end_row, start_row + max_rows - 1)

    rows: List[Dict[str, Any]] = []
    header_by_index = {int(c["index"]): c for c in columns}
    for row_idx in range(start_row, end_row + 1):
        values: Dict[str, Any] = {}
        has_text = False
        for col_idx in indices:
            value = normalize_text(ws.cell(row=row_idx, column=col_idx).value)
            if value:
                has_text = True
            col = header_by_index.get(col_idx, {})
            key = f"{get_column_letter(col_idx)}:{col.get('header', '')}"
            values[key] = truncate(value)
        if has_text:
            rows.append({"row": row_idx, "values": values})
    return rows


def workbook_outline(path: Path, max_rows: int, header_scan_rows: int) -> Dict[str, Any]:
    wb = load_workbook(path, read_only=False, data_only=True)
    output: Dict[str, Any] = {
        "workbook": str(path),
        "sheet_count": len(wb.sheetnames),
        "sheets": [],
    }
    for ws in wb.worksheets:
        candidates = candidate_header_rows(ws, header_scan_rows)
        selected_header = candidates[0]["row"] if candidates else None
        columns = detect_columns(ws, selected_header)
        sheet_info = {
            "name": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "candidate_header_rows": candidates,
            "selected_header_row": selected_header,
            "detected_columns": columns,
            "merged_ranges_sample": [str(rng) for rng in list(ws.merged_cells.ranges)[:30]],
            "styled_or_commented_cells_sample": extract_styled_or_commented_cells(ws),
            "text_rows": extract_text_rows(ws, selected_header, columns, max_rows),
        }
        output["sheets"].append(sheet_info)
    return output


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Extract a read-only outline from an Excel workbook.")
    parser.add_argument("workbook", type=Path, help="Path to .xlsx workbook")
    parser.add_argument("--out", type=Path, default=None, help="Output JSON path. Defaults to stdout.")
    parser.add_argument("--max-rows", type=int, default=5000, help="Max rows per sheet to include after header row. Use 0 for all rows.")
    parser.add_argument("--header-scan-rows", type=int, default=DEFAULT_HEADER_SCAN_ROWS, help="Number of top rows to scan for headers.")
    args = parser.parse_args(argv)

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    outline = workbook_outline(args.workbook, args.max_rows, args.header_scan_rows)
    text = json.dumps(outline, ensure_ascii=False, indent=2)
    if args.out:
        args.out.write_text(text, encoding="utf-8")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
